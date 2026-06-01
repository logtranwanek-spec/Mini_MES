from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from pyxlsb import open_workbook
from openpyxl import Workbook
import io
import os
import tempfile
import traceback

app = Flask(__name__)
CORS(app)

@app.route('/convert', methods=['POST'])
def convert_xlsb_to_xlsx():
    """Convert XLSB to XLSX and return file"""
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    # Lấy danh sách các sheet cần convert (nếu có)
    # Ví dụ: "Print,4,13"
    target_sheets_str = request.form.get('sheets', '')
    target_sheets = [s.strip() for s in target_sheets_str.split(',')] if target_sheets_str else []
    
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    
    if not file.filename.lower().endswith('.xlsb'):
        return jsonify({'error': 'Only XLSB files allowed'}), 400
    
    temp_xlsb_path = None
    
    try:
        print(f"📥 Received: {file.filename}")
        if target_sheets:
            print(f"   🎯 Target sheets only: {target_sheets}")
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsb') as tmp_xlsb:
            file.save(tmp_xlsb.name)
            temp_xlsb_path = tmp_xlsb.name
        
        xlsx_buffer = io.BytesIO()
        wb_out = Workbook()
        wb_out.remove(wb_out.active)
        
        with open_workbook(temp_xlsb_path) as wb_in:
            sheet_names = wb_in.sheets
            
            for sheet_name in sheet_names:
                # 🚀 TỐI ƯU: Bỏ qua các sheet không cần thiết
                if target_sheets and sheet_name not in target_sheets:
                    print(f"  ⏩ Skipping sheet: {sheet_name}")
                    continue
                    
                print(f"  📄 Processing sheet: {sheet_name}")
                
                try:
                    ws_out = wb_out.create_sheet(title=sheet_name)
                    
                    with wb_in.get_sheet(sheet_name) as sheet:
                        row_count = 0
                        for row in sheet.rows():
                            for cell in row:
                                if cell.v is not None:
                                    try:
                                        # 🚀 ĐÃ SỬA: Ép dùng tọa độ tuyệt đối của Excel (cell.r và cell.c)
                                        ws_out.cell(row=cell.r + 1, column=cell.c + 1, value=cell.v)
                                    except:
                                        pass
                            if row:
                                row_count = row[0].r + 1

                        print(f"    ✅ Sheet '{sheet_name}': {row_count} rows")
                
                except Exception as sheet_error:
                    print(f"    ❌ Error processing sheet '{sheet_name}': {sheet_error}")
                    continue
        
        wb_out.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        
        if temp_xlsb_path and os.path.exists(temp_xlsb_path):
            try: os.unlink(temp_xlsb_path)
            except: pass
        
        output_filename = file.filename.replace('.xlsb', '.xlsx')
        print(f"✅ Converted successfully!")
        
        return send_file(
            xlsx_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_filename
        )
        
    except Exception as e:
        if temp_xlsb_path and os.path.exists(temp_xlsb_path):
            try: os.unlink(temp_xlsb_path)
            except: pass
            
        print(f"❌ Error: {str(e)}")
        return jsonify({'error': str(e)}), 500


@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'ok',
        'service': 'XLSB to XLSX Converter',
        'version': '1.0.0'
    })

@app.route('/', methods=['GET'])
def index():
    """Root endpoint - service info"""
    return jsonify({
        'service': 'XLSB to XLSX Converter Service',
        'version': '1.0.0',
        'endpoints': {
            '/convert': 'POST - Convert XLSB to XLSX',
            '/health': 'GET - Health check'
        },
        'status': 'running'
    })

if __name__ == '__main__':
    print("=" * 70)
    print("  XLSB to XLSX Converter Service")
    print("  Version: 1.0.0")
    print("  Running on: http://localhost:5001")
    print("=" * 70)
    print()
    print("📋 Available endpoints:")
    print("   POST /convert - Convert XLSB file to XLSX")
    print("   GET  /health  - Health check")
    print("   GET  /        - Service info")
    print()
    print("⚡ Ready to receive requests...")
    print("=" * 70)
    
    app.run(host='0.0.0.0', port=5001, debug=False)